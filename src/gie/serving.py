"""Serving queries for the viewer's API.

All reads go through DuckDB over blob (cloud-optimized). The FastAPI layer
(api/) turns these into GeoJSON/JSON for the deck.gl + MapLibre front end.
"""

from __future__ import annotations

import geopandas as gpd
import numpy as np
import pandas as pd

from gie import db
from gie.config import load_settings


def _gold(settings, source: str, adm0: str) -> str:
    return settings.az_path("gold", f"source={source}", f"adm0={adm0}", "damage_facts.parquet")


def load_h3_damage(
    source: str = "microsoft", adm0: str = "VE", stage: str = "dev"
) -> pd.DataFrame:
    """Per-H3-cell damage metrics (wide), for the hexagon layer."""
    settings = load_settings(stage)  # type: ignore[arg-type]
    con = db.connect()
    gold = _gold(settings, source, adm0)
    return con.execute(
        f"""
        SELECT unit_id AS h3,
            max(value) FILTER (WHERE metric='buildings_total')  AS buildings_total,
            max(value) FILTER (WHERE metric='buildings_damaged') AS buildings_damaged,
            max(value) FILTER (WHERE metric='damaged_fraction')  AS damaged_fraction,
            max(value) FILTER (WHERE metric='damage_pct_mean')   AS damage_pct_mean
        FROM read_parquet('{gold}') WHERE unit_type='h3' GROUP BY unit_id
        """
    ).df()


def load_admin_damage(
    level: int = 3, source: str = "microsoft", adm0: str = "VE", stage: str = "dev"
) -> gpd.GeoDataFrame:
    """Admin units at ``level`` joined to damage facts.

    For adm3 we return every parroquia within the affected municipalities
    (siblings render as 'no data' for context); for adm1/adm2, only affected
    units. Result carries geometry + metrics for a choropleth.
    """
    settings = load_settings(stage)  # type: ignore[arg-type]
    con = db.connect()
    gold = _gold(settings, source, adm0)
    adm = settings.az_path("bronze", "source=codab", f"adm0={adm0}", f"adm{level}.parquet")
    idcol, namecol = f"adm{level}_id", f"adm{level}_name"

    if level >= 3:
        # parroquias within affected municipalities (context around the data)
        where = (
            f"a.adm2_id IN (SELECT DISTINCT unit_id "
            f"FROM read_parquet('{gold}') WHERE unit_type='adm2')"
        )
    else:
        where = "f.buildings_total IS NOT NULL"  # affected units only

    df = con.execute(
        f"""
        WITH facts AS (
            SELECT unit_id AS {idcol},
                max(value) FILTER (WHERE metric='buildings_total')   AS buildings_total,
                max(value) FILTER (WHERE metric='buildings_damaged') AS buildings_damaged,
                max(value) FILTER (WHERE metric='damaged_fraction')  AS damaged_fraction
            FROM read_parquet('{gold}') WHERE unit_type='adm{level}' GROUP BY unit_id
        )
        SELECT a.{idcol} AS unit_id, a.{namecol} AS unit_name, ST_AsWKB(a.geometry) AS wkb,
               f.buildings_total, f.buildings_damaged, f.damaged_fraction
        FROM read_parquet('{adm}') a
        LEFT JOIN facts f USING ({idcol})
        WHERE {where}
        """
    ).df()
    geom = gpd.GeoSeries.from_wkb(df.pop("wkb").map(bytes), crs="EPSG:4326")
    return gpd.GeoDataFrame(df, geometry=geom, crs="EPSG:4326")


def load_footprints(
    source: str = "microsoft", adm0: str = "VE", stage: str = "dev"
) -> gpd.GeoDataFrame:
    """Raw building footprints with damage attributes, for the footprint layer."""
    settings = load_settings(stage)  # type: ignore[arg-type]
    con = db.connect()
    silver = settings.az_path("silver", f"source={source}", f"adm0={adm0}", "footprints.parquet")
    df = con.execute(
        f"""
        SELECT damaged, damage_pct_10m, ST_AsWKB(geometry) AS wkb
        FROM read_parquet('{silver}')
        """
    ).df()
    geom = gpd.GeoSeries.from_wkb(df.pop("wkb").map(bytes), crs="EPSG:4326")
    return gpd.GeoDataFrame(df, geometry=geom, crs="EPSG:4326")


# --- common-model (gold/model=common) readers ------------------------------
_COMMON_PIVOT = """
    max(value) FILTER (WHERE metric='exposed_buildings')     AS exposed_buildings,
    max(value) FILTER (WHERE metric='analysed_buildings')    AS analysed_buildings,
    max(value) FILTER (WHERE metric='coverage_fraction')     AS coverage_fraction,
    max(value) FILTER (WHERE metric='damaged_detected')      AS damaged_detected,
    max(value) FILTER (WHERE metric='damaged_extrapolated')  AS damaged_extrapolated,
    max(value) FILTER (WHERE metric='analysed_area_km2')     AS analysed_area_km2,
    max(value) FILTER (WHERE metric='unit_area_km2')         AS unit_area_km2,
    max(value) FILTER (WHERE metric='area_coverage_fraction') AS area_coverage_fraction,
    -- CEMS-only hover breakdown (null for Microsoft): snapped grade counts + coarse estimate
    max(value) FILTER (WHERE metric='cems_destroyed')        AS cems_destroyed,
    max(value) FILTER (WHERE metric='cems_damaged')          AS cems_damaged,
    max(value) FILTER (WHERE metric='cems_possibly')         AS cems_possibly,
    max(value) FILTER (WHERE metric='cems_coarse_detected')  AS cems_coarse_detected
"""


def list_sources(adm0: str = "VE", stage: str = "dev") -> list[str]:
    """Distinct damage sources present in the common-model gold."""
    settings = load_settings(stage)  # type: ignore[arg-type]
    con = db.connect()
    gold = settings.az_path("gold", "model=common", f"adm0={adm0}", "facts.parquet")
    rows = con.execute(
        f"SELECT DISTINCT source FROM read_parquet('{gold}') ORDER BY source"
    ).fetchall()
    return [r[0] for r in rows]


def load_common_h3(source: str, adm0: str = "VE", stage: str = "dev") -> pd.DataFrame:
    """Per-H3-cell common-model metrics for one source."""
    settings = load_settings(stage)  # type: ignore[arg-type]
    con = db.connect()
    gold = settings.az_path("gold", "model=common", f"adm0={adm0}", "facts.parquet")
    return con.execute(
        f"SELECT unit_id AS h3, {_COMMON_PIVOT} FROM read_parquet('{gold}') "
        f"WHERE unit_type='h3' AND source='{source}' GROUP BY unit_id"
    ).df()


def load_common_admin(
    level: int, source: str, adm0: str = "VE", stage: str = "dev"
) -> gpd.GeoDataFrame:
    """Admin units (with geometry) joined to one source's common-model metrics."""
    settings = load_settings(stage)  # type: ignore[arg-type]
    con = db.connect()
    gold = settings.az_path("gold", "model=common", f"adm0={adm0}", "facts.parquet")
    adm = settings.az_path("bronze", "source=codab", f"adm0={adm0}", f"adm{level}.parquet")
    idcol = f"adm{level}_id"
    df = con.execute(
        f"""
        WITH facts AS (
            SELECT unit_id AS {idcol}, {_COMMON_PIVOT}
            FROM read_parquet('{gold}')
            WHERE unit_type='adm{level}' AND source='{source}' GROUP BY unit_id
        )
        SELECT a.{idcol} AS unit_id, a.adm{level}_name AS unit_name,
               ST_AsWKB(a.geometry) AS wkb, f.* EXCLUDE ({idcol})
        FROM read_parquet('{adm}') a JOIN facts f USING ({idcol})
        """
    ).df()
    geom = gpd.GeoSeries.from_wkb(df.pop("wkb").map(bytes), crs="EPSG:4326")
    return gpd.GeoDataFrame(df, geometry=geom, crs="EPSG:4326")


def load_export(level: int, adm0: str = "VE", stage: str = "dev") -> pd.DataFrame:
    """Tidy per-admin-unit, per-source damage table for spreadsheet export.

    One row per (admin unit x source): building counts, damage fraction (damaged
    / analysed valid-area buildings), and areal coverage. Full COD name hierarchy
    for context. Sorted so a unit's sources sit on adjacent rows for comparison.
    """
    settings = load_settings(stage)  # type: ignore[arg-type]
    con = db.connect()
    gold = settings.az_path("gold", "model=common", f"adm0={adm0}", "facts.parquet")
    adm = settings.az_path("bronze", "source=codab", f"adm0={adm0}", f"adm{level}.parquet")
    names = ", ".join(f"a.adm{i}_name" for i in range(1, level + 1))
    return con.execute(
        f"""
        WITH f AS (
            SELECT unit_id, source,
                max(value) FILTER (WHERE metric='exposed_buildings')      AS total_buildings,
                max(value) FILTER (WHERE metric='analysed_buildings')     AS analysed_buildings,
                max(value) FILTER (WHERE metric='damaged_detected')       AS damaged,
                max(value) FILTER (WHERE metric='analysed_area_km2')      AS analysed_area_km2,
                max(value) FILTER (WHERE metric='unit_area_km2')          AS unit_area_km2,
                max(value) FILTER (WHERE metric='area_coverage_fraction') AS area_coverage_fraction
            FROM read_parquet('{gold}') WHERE unit_type='adm{level}' GROUP BY unit_id, source
        )
        SELECT {names}, f.unit_id, f.source,
               f.total_buildings, f.analysed_buildings,
               f.analysed_buildings / nullif(f.total_buildings, 0) AS pct_buildings_covered,
               f.damaged,
               f.damaged / nullif(f.analysed_buildings, 0) AS damage_fraction,
               f.analysed_area_km2, f.unit_area_km2, f.area_coverage_fraction
        FROM f JOIN read_parquet('{adm}') a ON a.adm{level}_id = f.unit_id
        ORDER BY {names}, f.source
        """
    ).df()


# column -> (meaning / how it was derived), documented in the export README sheet
_EXPORT_GLOSSARY = [
    (
        "adm1_name / adm2_name / adm3_name",
        "OCHA COD administrative unit names — the hierarchy this row belongs to.",
    ),
    ("unit_id", "OCHA COD pcode of the admin unit."),
    (
        "source",
        "Damage source: 'microsoft' (AI per-building damage on Microsoft footprints) or "
        "'copernicus_ems' (Copernicus EMS rapid-mapping damage grades).",
    ),
    (
        "total_buildings",
        "Total buildings in the unit: Overture footprints whose centroid is inside it "
        "(Overture = Microsoft ML + Google Open Buildings + OSM, deduplicated; pulled for "
        "the full admin-1 state, so the count is complete). Identical for both sources.",
    ),
    (
        "analysed_buildings",
        "Of total_buildings, those inside the source's valid analysed area — where it "
        "actually assessed. Copernicus: image footprint within the AOI, minus cloud / "
        "no-data. Microsoft: its valid-area mask.",
    ),
    (
        "pct_buildings_covered",
        "analysed_buildings / total_buildings — share of the unit's buildings the source "
        "was able to assess.",
    ),
    (
        "damaged",
        "Buildings flagged damaged within the analysed area, counting any grade together "
        "(Possibly damaged, Damaged, or Destroyed). Microsoft: an AI-flagged damaged "
        "footprint. Copernicus EMS: an Overture building matched to a per-building damage "
        "point from its latest assessment (points snapped to the nearest footprint within "
        "20 m). Where a Copernicus AOI has only its earlier coarse damage-area blocks and "
        "no points yet, those blocks are used instead.",
    ),
    (
        "damage_fraction",
        "damaged / analysed_buildings — the observed damage rate within the assessed area "
        "(NOT over the whole unit).",
    ),
    (
        "analysed_area_km2",
        "Area (km²) of the source's analysed extent that falls within the unit.",
    ),
    ("unit_area_km2", "The admin unit's own area (km²)."),
    (
        "area_coverage_fraction",
        "analysed_area_km2 / unit_area_km2 — share of the unit's land area the source imaged.",
    ),
]
def export_workbook(adm0: str = "VE", stage: str = "dev") -> bytes:
    """Styled workbook: a README (with column glossary) + adm1/2/3 data sheets.
    Header styling mirrors the team's storm-exposure export."""
    import io
    from datetime import date

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor="55B284")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    band_fill = PatternFill("solid", fgColor="EAF4EE")  # zebra rows
    hair = Side(style="thin", color="D3D3D3")
    border = Border(left=hair, right=hair, top=hair, bottom=hair)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    top = Alignment(vertical="top", wrap_text=True)

    wb = Workbook()
    rm = wb.active
    rm.title = "README"
    rm.sheet_view.showGridLines = False
    rm["A1"] = "Venezuela Earthquake — Building Damage Exposure by Admin Unit"
    rm["A1"].font = Font(bold=True, size=22, color="3E8F6B")
    rm["A2"] = "Microsoft vs Copernicus EMS — buildings & damage by OCHA COD admin 1 / 2 / 3"
    rm["A2"].font = Font(italic=True, size=13, color="333333")
    rm["A3"] = (
        f"OCHA Centre for Humanitarian Data  ·  activation EMSR884  ·  "
        f"generated {date.today().isoformat()}  ·  dev"
    )
    rm["A3"].font = Font(size=10, color="888888")
    # teal section bar (spans columns A:B)
    rm["A5"] = "Columns — meaning & derivation"
    rm["A5"].font = Font(bold=True, size=12, color="FFFFFF")
    rm["A5"].fill = rm["B5"].fill = header_fill
    # glossary table header
    rm["A6"], rm["B6"] = "Column", "How it is derived"
    for cell in (rm["A6"], rm["B6"]):
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", indent=1)
        cell.border = border
    for i, (col, desc) in enumerate(_EXPORT_GLOSSARY, 7):
        rm[f"A{i}"], rm[f"B{i}"] = col, desc
        rm[f"A{i}"].font = Font(bold=True, color="333333")
        rm[f"B{i}"].font = Font(color="333333")
        for cell in (rm[f"A{i}"], rm[f"B{i}"]):
            cell.alignment = top
            cell.border = border
            if i % 2 == 0:
                cell.fill = band_fill
        rm.row_dimensions[i].height = max(16, 15 * (len(desc) // 88 + 1))
    rm.column_dimensions["A"].width = 30
    rm.column_dimensions["B"].width = 92

    pct = {"pct_buildings_covered", "damage_fraction", "area_coverage_fraction"}
    for level in (1, 2, 3):
        df = load_export(level, adm0, stage)
        ws = wb.create_sheet(f"adm{level}")
        ws.append(list(df.columns))
        for _, row in df.iterrows():
            ws.append([None if pd.isna(v) else v for v in row.tolist()])
        for cell in ws[1]:
            cell.fill, cell.font = header_fill, header_font
            cell.alignment, cell.border = center, border
        ws.row_dimensions[1].height = 30
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        ws.sheet_view.showGridLines = False
        for i, name in enumerate(df.columns, 1):
            letter = get_column_letter(i)
            ws.column_dimensions[letter].width = max(
                len(name) + 2, 22 if name in ("unit_id", "source") else 14
            )
            fmt = "0.0%" if name in pct else ("0.00" if name.endswith("km2") else "#,##0")
            if name not in ("source",) and "name" not in name and name != "unit_id":
                for cell in ws[letter][1:]:
                    cell.number_format = fmt
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            if row[0].row % 2 == 0:
                for cell in row:
                    cell.fill = band_fill

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def load_buildings(source: str, adm0: str = "VE", stage: str = "dev") -> pd.DataFrame:
    """Assessed buildings (as points) for one source, with a damaged flag.

    Reads the persisted per-building flags directly (deduped, assessed-only, with
    lon/lat) — no scan of the multi-million-row Overture base at serve time.
    """
    settings = load_settings(stage)  # type: ignore[arg-type]
    con = db.connect()
    flags = settings.az_path("gold", "model=common", f"adm0={adm0}", "building_flags.parquet")
    dmg = "ms_dmg" if source == "microsoft" else "cems_dmg"
    seen = "ms_analysed" if source == "microsoft" else "cems_analysed"
    return con.execute(
        f"SELECT lon, lat, {dmg}::INT AS damaged FROM read_parquet('{flags}') WHERE {seen}"
    ).df()


def load_source_extent(source: str, adm0: str = "VE", stage: str = "dev") -> gpd.GeoDataFrame:
    """Each source's real analysis extent as labelling outlines, one per AOI /
    product (kept separate, not dissolved), carrying metadata for hover.

    Microsoft -> its valid-area masks (one per AOI). CEMS -> one outline per
    product (the actual analysed swath, dissolved only within a product), with
    AOI name, product kind (initial vs monitoring) and acquisition date. NOT a
    convex hull, which would inflate the narrow swaths and overstate coverage.
    """
    settings = load_settings(stage)  # type: ignore[arg-type]
    con = db.connect()
    src = "microsoft" if source == "microsoft" else "copernicus_ems"
    ext = settings.az_path("silver", f"source={src}", f"adm0={adm0}", "analysed_extent.parquet")
    if source == "microsoft":
        sql = (
            f"SELECT aoi AS aoi_name, 'Microsoft analysis' AS product, NULL AS acquired, "
            f"ST_AsWKB(geometry) AS wkb FROM read_parquet('{ext}')"
        )
    else:
        sql = (
            f"SELECT any_value(aoi_name) AS aoi_name, any_value(product) AS product, "
            f"any_value(acquired) AS acquired, "
            f"ST_AsWKB(ST_Union_Agg(ST_MakeValid(geometry))) AS wkb "
            f"FROM read_parquet('{ext}') GROUP BY src_zip"
        )
    df = con.execute(sql).df()
    geom = gpd.GeoSeries.from_wkb(df.pop("wkb").map(bytes), crs="EPSG:4326")
    df["source"] = source
    return gpd.GeoDataFrame(df, geometry=geom, crs="EPSG:4326")


def load_native(source: str, adm0: str = "VE", stage: str = "dev") -> gpd.GeoDataFrame:
    """Each source's own damage geometry (the 'native data' view).

    Microsoft -> footprints with a binary damaged flag; CEMS -> damage-grade
    polygons with the EMS grade. Distinct geometries that overlay cleanly and
    show what each source actually mapped, vs the Overture-base view.
    """
    settings = load_settings(stage)  # type: ignore[arg-type]
    con = db.connect()
    if source == "microsoft":
        path = settings.az_path("silver", "source=microsoft", f"adm0={adm0}", "footprints.parquet")
        cols = "damaged"
    else:
        path = settings.az_path(
            "silver", "source=copernicus_ems", f"adm0={adm0}", "builtup_damage.parquet"
        )
        cols = "ems_grade, damage_class"
    df = con.execute(
        f"SELECT {cols}, ST_AsWKB(geometry) AS wkb FROM read_parquet('{path}')"
    ).df()
    geom = gpd.GeoSeries.from_wkb(df.pop("wkb").map(bytes), crs="EPSG:4326")
    return gpd.GeoDataFrame(df, geometry=geom, crs="EPSG:4326")


def load_coverage_detail(adm0: str = "VE", stage: str = "dev") -> gpd.GeoDataFrame:
    """CEMS Area-of-Interest + Not-Analysed (cloud) shapes, for the native view.

    Shows what the assessment looked at (AOI) and the holes within it that
    couldn't be assessed (cloud / no imagery) — so coverage gaps are visible.
    """
    settings = load_settings(stage)  # type: ignore[arg-type]
    con = db.connect()
    path = settings.az_path(
        "silver", "source=copernicus_ems", f"adm0={adm0}", "coverage_detail.parquet"
    )
    df = con.execute(
        f"SELECT kind, aoi_name, product, acquired, ST_AsWKB(geometry) AS wkb "
        f"FROM read_parquet('{path}')"
    ).df()
    geom = gpd.GeoSeries.from_wkb(df.pop("wkb").map(bytes), crs="EPSG:4326")
    return gpd.GeoDataFrame(df, geometry=geom, crs="EPSG:4326")


def load_agreement(adm0: str = "VE", stage: str = "dev") -> pd.DataFrame:
    """Per-building source-agreement category (MS vs CEMS) for the agreement map.

    Where BOTH sources assessed a building (the overlap): both / ms_only /
    cems_only / agree_none. Where only one looked: ms_area / cems_area. This is
    the spatial Venn — it shows where the sources agree and, crucially, where
    they disagree.
    """
    settings = load_settings(stage)  # type: ignore[arg-type]
    con = db.connect()
    flags = settings.az_path("gold", "model=common", f"adm0={adm0}", "building_flags.parquet")
    return con.execute(
        f"""
        SELECT lon, lat,
               CASE
                 WHEN ms_analysed AND cems_analysed THEN
                   CASE WHEN ms_dmg AND cems_dmg THEN 'both'
                        WHEN ms_dmg THEN 'ms_only'
                        WHEN cems_dmg THEN 'cems_only'
                        ELSE 'agree_none' END
                 WHEN ms_analysed THEN 'ms_area'
                 ELSE 'cems_area' END AS agreement
        FROM read_parquet('{flags}')
        """
    ).df()


def damage_colors(
    fractions, *, na: tuple[int, int, int, int] = (200, 200, 200, 40)
) -> np.ndarray:
    """Map a damaged-fraction series (0..1, NaN allowed) to an RGBA uint8 array."""
    f = np.asarray(fractions, dtype="float64")
    out = np.empty((len(f), 4), dtype="uint8")
    valid = ~np.isnan(f)
    fc = np.clip(f[valid], 0.0, 1.0)
    out[valid, 0] = 240
    out[valid, 1] = (220 * (1 - fc)).astype("uint8")
    out[valid, 2] = (40 * (1 - fc)).astype("uint8")
    out[valid, 3] = 200
    out[~valid] = na
    return out
